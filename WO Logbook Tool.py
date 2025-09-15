#WO Logbook Tool
import os, re, tkinterweb, json, sys
import pyodbc
from openpyxl import load_workbook, Workbook
import datetime
import time
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

# Add the parent directory ('NCC Automations') to the Python path
# This allows us to import the 'PythonTools' package from there.
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(parent_dir)
from PythonTools import CREDS, EMAILS, LOGBOOK_LOCATION_MAP, CUSTOMERS_SITES_EMAINT, CUSTOMERS_SITES_NORMAL_NAMING, restart_pc, legible_date_validation, time_validation


slr_customer_data = []
solt_customer_data = []
nar_customer_data = []
nce_customer_data = []
hst_customer_data = []

xelio_customer_data = []
charter_customer_data = []

#Tuple is as follows = name, wo-report, access_log_report 
#Will be used to identify users who get the report or do not. 
customer_report_items = [('Harrison St.', hst_customer_data, False), ('NARENCO', nar_customer_data, False), 
                         ('NCEMC', nce_customer_data, False), 
                         ('Sol River', slr_customer_data, False), ('Soltage', solt_customer_data, True)]


def dbcnxn():
    global db, connect_db, c
    #Connect to DB
    realdb = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=G:\Shared drives\Narenco Projects\O&M Projects\NCC\NCC\NCC 039.accdb;'
    db = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=G:\Shared drives\Narenco Projects\O&M Projects\NCC\NCC\NCC 039 - Copy.accdb'
    if testingvar.get():
        connect_db = pyodbc.connect(db)
    else:    
        connect_db = pyodbc.connect(realdb)
    c = connect_db.cursor()

def browse_files():
    # Get the path to the Downloads folder
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    # Open file dialog to select the first file
    wo_data = filedialog.askopenfilename(
        title="Select the Emaint Report Download Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialdir=downloads_folder
    )
    if not wo_data:
        return
    else:
        parse_wo(wo_data)

#Shift SUMMARY Functions
def shift_Summary():
    db = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=G:\Shared drives\Narenco Projects\O&M Projects\NCC\NCC\NCC 039.accdb;'
    connect_dbn = pyodbc.connect(db)
    c = connect_dbn.cursor()


    # MS Access SQL uses Date() for the current date and DateValue() to extract the date part from a DateTime field.
    query_min_id = """
        SELECT MIN(ActivityLogID)
        FROM [ShiftSummary]
        WHERE [EditDate] IS NOT NULL
          AND YEAR([EditDate]) = YEAR(Date())
          AND MONTH([EditDate]) = MONTH(Date())
          AND DAY([EditDate]) = DAY(Date())
    """
    initial_activity_id = None
    try:
        c.execute(query_min_id)
        result = c.fetchone()
        if result and result[0] is not None:
            initial_activity_id = int(result[0])
        else:
            print("No entries found for today in ShiftSummary, or MIN(ActivityLogID) was NULL.")
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        messagebox.showerror("Database Error", f"Error fetching initial ActivityLogID: {sqlstate}\nPlease check the table/column names and database connection.\n{ex}", parent=root)
        # initial_activity_id will remain None, dialog will show without a pre-filled value or user can cancel.

    # --- Modified simpledialog call ---
    firstentry = simpledialog.askinteger(
        title="Shift Summary Query",
        prompt="Confirm or input the ActivityLogID for today's first entry:",
        initialvalue=initial_activity_id,  # Set the initial value here
        parent=root # Good practice to set parent for dialogs
    )


    c.execute(f"SELECT * FROM [ShiftSummary] WHERE [ActivityLogID] >= {firstentry}")
    todays_entries = c.fetchall()

    todays_date = datetime.datetime.today().strftime('%m/%d/%Y')

    # Initialize a dictionary to accumulate data for each ActivityLogID
    activity_data = {}

    # Iterate over each row in the entries
    for row in todays_entries:
        activitylog_id = row[0]
        employee = row[5]
        dickey = ''.join((str(activitylog_id), str(employee)))

        # Check if the current ActivityLogID has already been processed
        if dickey not in activity_data:
            if row[8]:
                worktype = str(row[7]) if row[7] else ''
                workLog = ' | '.join((worktype, str(row[8]) if row[8] else ''))
            else:
                workLog = ''
            if row[3] and row[4] and row[5] and row[6]:
                siteAccess = ' | '.join(('', str(row[3]), str(row[4]), str(row[5]), str(row[6])))
            else:
                siteAccess = ''
            if row[13]:
                issueLog = ' | '.join((str(row[9]) if row[9] else '', str(row[10]) if row[10] else '', str(row[11]) if row[11] else '', str(row[12]) if row[12] else '', str(row[13]) if row[13] else ''))
            else:
                issueLog = ''
            
            try:
                starttime = ' '.join((str(row[14].date()), str(row[15].time())))
            except AttributeError:
                starttime = ''
                messagebox.showwarning(title="Missing Start Date/Time", message=f"Log Id: {activitylog_id}\nMissing Start Date/Time. Field will be blank in Shift Summary Table. Please input a Start Date/Time for {activitylog_id} and refresh table with 'Update Table' Button on Preview Window")
            if row[16] and row[17]:
                endtime = ' '.join((str(row[16].date()), str(row[17].time())))
            else:
                endtime = ''
            # Initialize a new entry in the dictionary
            activity_data[dickey] = {
                'location': row[1],
                'activity': row[2],
                'site_access': siteAccess,
                'site_work': [workLog],
                'issue_tracking': issueLog,
                'start_datetime': starttime,
                'end_datetime': endtime,
            }
        else:
            # Add data from row[7] and row[8] to the notes list
            if row[7] is not None and row[8] is not None:
                workLog = ' | '.join((str(row[7]), str(row[8])))
                activity_data[dickey]['site_work'].append(workLog)


    # Construct the HTML table
    html_table = f"<h2 style='text-align: center; color: black;'>NCC Daily Activity Report for {todays_date}</h2>"
    html_table += "<table style='border-collapse: collapse; width: 100%;'>"
    html_table += "<tr style='background-color: lightgray;'>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Location</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Activity</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Site Access</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Site Work</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Issue Tracking</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Start Date/Time</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>End Date/Time</th>"
    html_table += "</tr>"

    # Iterate over the accumulated data to build the HTML rows
    for dickey, data in activity_data.items():
        html_table += "<tr>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{data['location']}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{data['activity']}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{data['site_access']}</td>"
        notes_combined = "<br>".join(data['site_work'])
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{notes_combined}</td>"

        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{data['issue_tracking']}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{data['start_datetime']}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{data['end_datetime']}</td>"

        html_table += "</tr>"

    html_table += "</table>"
    # Display the table in a new window
    preview_window = tk.Toplevel(root)
    preview_window.title("Shift Summary Preview")
    try:
        preview_window.iconbitmap(r"G:\Shared drives\O&M\NCC Automations\Icons\email-1.ico")
    except Exception as e:
        print(f"Error loading icon: {e}")
    preview_window.attributes("-fullscreen", True)

    # Create a canvas and a scrollbar
    canvas = tk.Canvas(preview_window)
    scrollbar = tk.Scrollbar(preview_window, orient=tk.VERTICAL, command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    # Pack the scrollable_frame to expand and fill
    scrollable_frame.pack(fill=tk.BOTH, expand=True)

    # Pack the canvas and scrollbar
    canvas.pack(fill=tk.BOTH, expand=True)

    # Create an HtmlFrame widget to display the HTML table
    html_frame = tkinterweb.HtmlFrame(scrollable_frame)
    html_frame.load_html(html_table)
    html_frame.pack(expand=True, fill=tk.BOTH)

    # Create buttons for restarting or sending the email
    button_frame = tk.Frame(preview_window)
    button_frame.pack(fill=tk.X)

    destroy_preview_button = tk.Button(button_frame, text="Close Preview", command=preview_window.destroy, width=50)
    destroy_preview_button.pack(side=tk.LEFT, padx=10, pady=10)

    destroy_root_button = tk.Button(button_frame, text="Close App", command=root.destroy, width=50)
    destroy_root_button.pack(side=tk.LEFT, padx=10, pady=10)

    restart_button = tk.Button(button_frame, text="Update Table", command=lambda: restart_shift_summary(preview_window), width=50)
    restart_button.pack(side=tk.RIGHT, padx=10, pady=10)

    send_button = tk.Button(button_frame, text="Send Email", command=lambda: send_shift_Summary(html_table, preview_window), width=50)
    send_button.pack(side=tk.RIGHT, padx=10, pady=10)

def restart_shift_summary(preview_window):
    preview_window.destroy()
    shift_Summary()

def send_shift_Summary(html_table, preview_window):
    date = datetime.datetime.now()
    today = date.strftime("%m/%d/%y")

    shift_sum_recipients = EMAILS['Shift Sum List']

    test = EMAILS['Joseph Lang']
    # Create the email message
    message = MIMEMultipart()
    message["Subject"] = f"{today} NCC Shift Summary"
    message["From"] = EMAILS['NCC Desk']
    message["To"] = ', '.join(shift_sum_recipients)
    #message["To"] = test
    password = CREDS['shiftsumEmail']
    sender = EMAILS['NCC Desk']
    
    # Attach HTML content
    html_body = MIMEText(html_table, "html")
    message.attach(html_body)

    # Send the email
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender, password)
        server.sendmail(sender, shift_sum_recipients, message.as_string())    
    
    time.sleep(2.5)
    restart_pc() #Force Restart

    
def parse_wo(wos):
    global c
    dbcnxn()

    activity_logbook_data = []
    wb = load_workbook(wos)
    ws = wb.active
    for index, row in enumerate(ws.iter_rows(values_only=True)):
        if index == 0:
            continue
        wo = None
        startdate = None
        starttime = None
        enddate = None
        endtime = None
        site = None
        activityid = None
        locationid = None
        userlistid = None
        issuelistid = None
        editDate = None
        activity_log_id = None
        wo_notes = None
        customer_note = None

        # Assuming row[10] contains HTML content
        clean_text = BeautifulSoup(row[8], "html.parser").get_text()

        if row[0] == "JOSEPHL99":
            userlistid = 7
        elif row[0] == "JACOBBUD99":
            userlistid = 10
        elif row[0] in {"NARENCO", "JAYMEORR99", "BRANDONA99", "BRANDONA96","NEWMANSE99"}:
            userlistid = 3
        else:
            messagebox.showerror(title="Exiting WO Input", message=f"Found Error in WO XL File\nUser is Neither:\nJoseph Lang\nJacob Budd\nBrandon Arrowood (Admin Account)\nWO: {wo}")
            root.destroy()

        wo = row[2]
        site = row[3]
        activityid = 3

        locationid = None
        for key, locations in LOGBOOK_LOCATION_MAP.items():
            if site in locations:
                locationid = key
                break
        if locationid is None:
            messagebox.showerror(title="Site Not Found", message=f"Site '{site}' not found in location dictionary.\nWO: {wo}")
            root.destroy()

        for customervar, siteList in CUSTOMERS_SITES_EMAINT.items():
            if site in siteList:
                customer = customervar
                break

        startdate_match = re.search(r'Start Date: (\d{1,2}[-/]\d{1,2}(?:[-/]\d{2,4})?)', clean_text)
        starttime_match = re.search(r'Start Time: (\d{1,2}[:;]?\d{1,2}|\d{3,4})', clean_text)
        enddate_match = re.search(r'End Date: (\d{1,2}[-/]\d{1,2}(?:[-/]\d{2,4})?)', clean_text)
        endtime_match = re.search(r'End Time: (\d{1,2}[:;]?\d{1,2}|\d{3,4})', clean_text)
        # Function to add current year if missing
        def add_current_year(date_str):
            # Check if the year is missing
            if re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', date_str):
                return date_str  # Year is present
            else:
                current_year = datetime.datetime.now().year
                return f"{date_str}/{current_year}"
        def format_time(time_str):
            # If the time is already in HH:MM format, return it as is
            if ':' in time_str:
                return time_str
            # If the time is a group of 3 or 4 digits, format it to HH:MM
            elif len(time_str) == 3:
                return f"0{time_str[0]}:{time_str[1:]}"
            elif len(time_str) == 4:
                return f"{time_str[:2]}:{time_str[2:]}"
            else:
                raise ValueError("Invalid time format")

        # Extract and format the start time
        if starttime_match:
            #print(starttime_match.group(1), clean_text)
            starttime = format_time(starttime_match.group(1))
            #print("Formatted Start Time:", starttime)

        # Extract and format the end time
        if endtime_match:
            endtime = format_time(endtime_match.group(1))
            #print("Formatted End Time:", endtime)

        # Extract the end date
        if enddate_match:
            enddate = add_current_year(enddate_match.group(1))
        # Extract the end date
        if startdate_match:
            startdate = add_current_year(startdate_match.group(1))
        editDate = datetime.date.today()
        
        # Insert data into the ActivityLog table and retrieve the ActivityLogID
        try:
            
            # Execute the insert statement
            c.execute("""
                INSERT INTO ActivityLog (ActivityListID, UserListID, LocationListID, StartDate, StartTime, EndDate, EndTime, EditDate)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (activityid, userlistid, locationid, startdate, starttime, enddate, endtime, editDate))
            
            # Retrieve the last inserted ActivityLogID
            c.execute("SELECT @@IDENTITY")
            activity_log_id = c.fetchone()[0]

            connect_db.commit()
            

        except Exception as e:
            messagebox.showerror(title="Database Error", message=f"An error occurred while inserting data into the database: {e}")
        
        stowmatch1 = re.search(r'stow', row[6], re.IGNORECASE)
        stowmatch2 = re.search(r'stow', row[7], re.IGNORECASE)
        stowmatch3 = re.search(r'stow', row[8], re.IGNORECASE)


        invnomatch = re.search(r'Inverter\s*(\d+)', row[6], re.IGNORECASE)
        if invnomatch:
            invno = invnomatch.group(1)
        else:
            invno = '0'

        invcheck1 = re.search(r'(Inverter|INV)', row[6], re.IGNORECASE)
        invcheck2 = re.search(r'(Inverter|INV)', row[7], re.IGNORECASE)

        underperformcheck1 = re.search(r'underperform', clean_text, re.IGNORECASE)
        underperformcheck2 = re.search(r'underperform', row[7], re.IGNORECASE)

        trcheck1 = re.search(r'(Node|TCU|Tracker)', row[6], re.IGNORECASE)
        trcheck2 = re.search(r'(Node|TCU|Tracker)', row[7], re.IGNORECASE)

        cbcheck1 = re.search(r'(CMB|CB|Combiner Box)', row[6], re.IGNORECASE)
        cbcheck2 = re.search(r'(CMB|CB|Combiner Box)', row[7], re.IGNORECASE)

        #Remote Repair Check
        rrcheck = re.search(r'Can we resolve the issue Remotely\? \(Y or N\) (Y|Yes|N|No|None)', clean_text, re.IGNORECASE)
        #print(rrcheck, clean_text)
        if rrcheck:
            response = rrcheck.group(1).lower()
            #print(response)
            if response in ['none', 'n', 'no']:
                repairs_note = "by the technician on site"
            else:
                repairs_note = 'by the NCC remotely'



        if row[4] == 'Site Outage' and (row[5] == 'Utility Trip' or row[5] == 'Utility Planned'):
            issuelistid = 165
            if enddate is None:
                customer_note = f'Utility Tripped the site offline at {starttime} on {startdate}'
            else:
                customer_note = f'Utility Tripped the site offline at {starttime} on {startdate}. Production restored to the site {repairs_note} at {endtime} on {enddate}'
        elif row[4] == 'Site Outage' and (row[5] == 'Site Trip' or row[5] == 'Nuisance Trip'):
            issuelistid = 167
            if enddate is None:
                customer_note = f'Site Tripped offline at {starttime} on {startdate}'
            else:
                customer_note = f'Site Tripped offline at {starttime} on {startdate}. Production restored to the site {repairs_note} at {endtime} on {enddate}'
        elif row[4] == 'Equipment Outage' and (invcheck1 is not None or invcheck2 is not None):
            issuelistid = 42
            if enddate is None:
                customer_note = f'Inverter {invno} tripped offline at {starttime} on {startdate}'
            else:
                customer_note = f'Inverter {invno} tripped offline at {starttime} on {startdate}. Production restored to the device {repairs_note} at {endtime} on {enddate}'
        elif row[4] == 'COMMs Outage' and (invcheck1 is not None or invcheck2 is not None):
            issuelistid = 64
        elif row[4] == 'Underperformance' and (underperformcheck1 is not None or underperformcheck2 is not None) and (invcheck1 is not None or invcheck2 is not None):
            issuelistid = 60
        elif row[4] == 'Underperformance' and (trcheck1 is not None or trcheck2 is not None):
            issuelistid = 229
        elif row[4] == 'Underperformance' and (cbcheck1 is not None or cbcheck2 is not None):
            issuelistid = 212
        elif stowmatch1 or stowmatch2 or stowmatch3:
            issuelistid = 227
            customer_note = 'Stowed for Local Thunderstorm or Excessive Wind Speeds'
        else:
            issuelistid = None

        wo_notes_match = re.search(r'Summary of issue:\s*(.*?)\s*Can we resolve the issue Remotely\?', clean_text, re.DOTALL)
        if wo_notes_match:
            wo_notes = f'WO {wo}  {wo_notes_match.group(1)}'
        else:
            wo_notes = f'WO {wo}'
        
        try:
            # Now you can use activity_log_id to insert more data into another table
            # Example: Insert into another table using activity_log_id
            
            c.execute("""
                INSERT INTO ObservationLog (ActivityLogID, DeviceListID, IssueListID, Notes)
                VALUES (?, ?, ?, ?)
            """, (activity_log_id, None, issuelistid, wo_notes))
            connect_db.commit()
            
        except Exception as e:
            messagebox.showerror(title="Database Error", message=f"An error occurred while inserting data into the database: {e}")
            root.destroy()
        

        #Customer Notification Process
        if customer_note is not None:
            globals()[f'{customer}_customer_data'].append([wo, site, startdate, starttime, enddate, endtime, customer_note]) 

    if messagebox.askokcancel(title="WO's to Logbook Success!",
    message= "Successfully input all WO's to the NCC Logbook\nClick OK to view Customer Reports\nClick Cancel to Close"):
        for customer, customer_data, site_access_log in customer_report_items:
            if site_access_log:
                site_access_table = site_access_query(customer)
                customer_noti(customer, customer_data, site_access_table)
            else:
                customer_noti(customer, customer_data, site_access_log)
    else:
        connect_db.close()


def send_email(customer_data, window, customer):
    #print(f"Custom Data: {customer_data}")
    # Create the email message
    today = datetime.date.today().strftime("%m/%d/%y")
    message = MIMEMultipart()
    me = EMAILS['Joseph Lang']
    brandon = EMAILS['Brandon Arrowood']
    if customer == 'Harrison St.':
        message["Subject"] = f"NARENCO O&M | {today} Incident Report - {customer}"
        recipients = EMAILS['Harrison St']
    elif customer == 'Sol River':
        message["Subject"] = f"NARENCO O&M | {today} Incident Report - {customer}"
        recipients = EMAILS['Sol River']
    elif customer == 'Soltage':
        message["Subject"] = f"NARENCO O&M | {today} Daily Report - {customer}"
        site_access = site_access_query('Soltage')
        message.attach(MIMEText(site_access, "html"))
        recipients = EMAILS['Soltage']
    elif customer == 'NCEMC':
        message["Subject"] = f"NARENCO O&M | {today} Incident Report - {customer}"
        recipients = EMAILS['NCEMC']
    elif customer == 'NARENCO':
        message["Subject"] = f"NARENCO O&M | {today} Incident Report - {customer}"
        recipients = EMAILS['NARENCO']
    

    message["From"] = EMAILS['NCC Desk']
    if testingvar.get():
        message["To"] = me
    else:
        message["To"] = ', '.join(recipients)

    password = CREDS['shiftsumEmail']
    sender = EMAILS['NCC Desk']

    # Title/Header
    email_table = f"<h2 style='text-align: center; color: black;'>NCC Daily WO Report - {customer}</h2>"
    # Initialize HTML table with border style and header row
    email_table += "<table style='border-collapse: collapse; width: 100%;'>"
    # Header row with lightgray background
    email_table += "<tr style='background-color: lightgray;'>"
    email_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Ticket #</th>"
    email_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Site</th>"
    email_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Start Date | Time</th>"
    email_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>End Date | Time</th>"
    email_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Issue</th>"
    email_table += "</tr>"
    
    # Iterate over each row in the entries
    for row in customer_data:
        # Start a new row
        email_table += "<tr>"
        email_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[0]}</td>"
        email_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[1]}</td>"
        email_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[2]} | {row[3]}</td>"
        if row[4] != 'None':
            email_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[4]} | {row[5]}</td>"
        else:
            email_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'> </td>"
        email_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{row[6]}</td>"
        # End the row
        email_table += "</tr>"
    # End the table
    email_table += "</table>"
    if customer_data:
        message.attach(MIMEText(email_table, "html"))

    # Send the email
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender, password)
        if testingvar.get():
            server.sendmail(sender, me, message.as_string())
        else:
            server.sendmail(sender, recipients, message.as_string())
    window.destroy()

def site_access_query(customer):
    global c
    dbcnxn()
    if customer == 'Soltage':
        sites = [7, 9, 22, 23, 24]
    #Add Other customers LocationIDs according to the NCC 039 Database
    #elif customer == '?'

    query = """
    SELECT al.StartDate, al.StartTime, al.EndDate, al.EndTime, a.PeopleCount, cl.Company, l.Location
    FROM ((AccessLog AS a
    INNER JOIN ActivityLog  AS al ON a.ActivityLogID = al.ActivityLogID)
    INNER JOIN ContactList AS cl ON a.ContactListID = cl.ContactListID)
    INNER JOIN LocationList as l ON al.LocationListID = l.LocationID
    WHERE al.LocationListID IN ({})
    AND al.StartDate = ?
    """.format(','.join('?' for _ in sites))
    
    today = datetime.date.today()  
    
    c.execute(query, *sites, today)
    activity_log_results = c.fetchall()

        # Title/Header
    html_table = "<h2 style='text-align: center; color: black;'>Access Log Report</h2>"
    # Initialize HTML table with border style and header row
    html_table += "<table style='border-collapse: collapse; width: 100%;'>"
    # Header row with lightgray background
    html_table += "<tr style='background-color: lightgray;'>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Location</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Company</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'># Personnel</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Start Time</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>End Time</th>"
    html_table += "</tr>"
    # Iterate over each row in the results
        # Combine rows with identical data except for PeopleCount
    combined_results = {}
    for row in activity_log_results:
        key = (row.StartDate, row.StartTime, row.EndDate, row.EndTime, row.Company, row.Location)
        if key in combined_results:
            combined_results[key] += row.PeopleCount
        else:
            combined_results[key] = row.PeopleCount

    for key, people_count in combined_results.items():
        start_date, start_time, end_date, end_time, company, location = key
        # Start a new row
        html_table += "<tr>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{location}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{company}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{people_count}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{start_time.strftime("%H:%M")}</td>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{end_time.strftime("%H:%M") if end_time else ""}</td>"
        # End the row
        html_table += "</tr>"
    
    if activity_log_results == []:
        html_table += "<tr>"
        html_table += f"<td style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;' colspan= '5'>No Site Access Today</td>"
        html_table += "</tr>"

    # End the table
    html_table += "</table>"
    connect_db.close()
    return html_table

def site_access_only():
    for customer, table, site_accessTF in customer_report_items:
        if site_accessTF:
            site_access_table = site_access_query(customer)
            customer_noti(customer, table, site_access_table)

def manual_wo_reports():
    for customer, customer_data, site_access_log in customer_report_items:
        if site_access_log:
            site_access_table = site_access_query(customer)
            customer_noti(customer, customer_data, site_access_table)
        else:
            customer_noti(customer, customer_data, site_access_log)

def customer_noti(customer, customer_data, site_access_table):
    # Title/Header
    html_table = f"<h2 style='text-align: center; color: black;'>NCC Daily WO Report - {customer}</h2>"
    # Initialize HTML table with border style and header row
    html_table += "<table style='border-collapse: collapse; width: 100%;'>"
    # Header row with lightgray background
    html_table += "<tr style='background-color: lightgray;'>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Ticket #</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Site</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Start Date | Time</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>End Date | Time</th>"
    html_table += "<th style='border: 1px solid black; padding: 8px; color: black;'>Issue</th>"
    html_table += "</tr>"
    
    # Iterate over each row in the entries
    for row in customer_data:
        # Start a new row
        html_table += "<tr>"
        html_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[0]}</td>"
        html_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[1]}</td>"
        html_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[2]} | {row[3]}</td>"
        if row[4]:
            html_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue; text-align: center;'>{row[4]} | {row[5]}</td>"
        else:
            html_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'> </td>"
        html_table += f"<td contenteditable='true' style='border: 1px solid black; padding: 8px; color: black; background-color: lightblue;'>{row[6]}</td>"
        # End the row
        html_table += "</tr>"
    # End the table
    html_table += "</table>"
    # Display the table in a new window
    preview_window = tk.Toplevel(root)
    preview_window.title(f"{customer} Report Preview")
    try:
        # Note: The icon path should be valid for this to work.
        # preview_window.iconbitmap(r"G:\Shared drives\O&M\NCC Automations\Icons\email-1.ico")
        pass # Pass for now to make it runnable
    except Exception as e:
        print(f"Error loading icon: {e}")
    preview_window.geometry("1000x800")
    preview_window.resizable(True, True)

    # Create a container frame for the canvas and scrollbar
    container_frame = tk.Frame(preview_window)
    container_frame.pack(fill="both", expand=True)

    # Create a canvas and a scrollbar
    canvas = tk.Canvas(container_frame)
    scrollbar = tk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollable_frame = tk.Frame(canvas)
    scrollable_frame_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    def on_frame_configure(event=None):
        """Reset the scroll region to encompass the inner frame"""
        canvas.configure(scrollregion=canvas.bbox("all"))

    scrollable_frame.bind("<Configure>", on_frame_configure)
    canvas.bind(
        "<Configure>",
        lambda e: FrameWidth(e, canvas, scrollable_frame_id)
    )
    
    def on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    canvas.bind_all("<MouseWheel>", on_mousewheel)


    if site_access_table:
        site_frame = tkinterweb.HtmlFrame(scrollable_frame)
        site_frame.load_html(site_access_table)
        site_frame.pack(fill="x", padx=10, pady=5)

    html_frame = tkinterweb.HtmlFrame(scrollable_frame)
    html_frame.load_html(html_table)
    html_frame.pack(fill="x", padx=10, pady=5)

    #Instructions
    main_lbl = tk.Label(scrollable_frame, text='Edit the Table above with the one Below', borderwidth=1, relief="solid", font=('TkDefaultFont', 16))
    main_lbl.pack(fill='x', padx=10, pady=(10,5))
    
    # Create a frame for the grid-based table
    grid_frame = tk.Frame(scrollable_frame)
    grid_frame.pack(fill='both', padx=10, pady=5)

    # Add headers to the grid
    headers = ["✓= Include", "WO", "Site", "Start Date", "Start Time", "End Date", "End Time", "Issue"]
    for col, header in enumerate(headers):
        label = tk.Label(grid_frame, text=header, borderwidth=1, relief="solid", padx=5, pady=5)
        label.grid(row=0, column=col, sticky="nsew")

    entry_widgets = []  # List to store entry widgets
    check_vars = []  # List to store checkbox variables

    # Populate the grid with existing data
    for row_index, row_data in enumerate(customer_data, start=1):
        entry_row = []
        check_var = tk.BooleanVar(value=True)
        check_button = tk.Checkbutton(grid_frame, variable=check_var)
        check_button.grid(row=row_index, column=0, sticky="nsew")
        check_vars.append(check_var)

        # The original data has 7 columns, so we map them to the 7 entry columns
        for col_index, value in enumerate(row_data):
            if col_index in [0, 1, 6]:
                entry = tk.Entry(grid_frame, borderwidth=1, relief="solid")
                entry.insert(0, str(value))
                entry.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(entry)
            elif col_index in [2, 4]:
                entry = tk.Entry(grid_frame, borderwidth=1, relief="solid", validate='all', validatecommand=vcmd_date)
                entry.insert(0, str(value))
                entry.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(entry)
            elif col_index in [3, 5]:
                entry = tk.Entry(grid_frame, borderwidth=1, relief="solid", validate='all', validatecommand=vcmd_time)
                entry.insert(0, str(value))
                entry.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(entry)
        entry_widgets.append(entry_row)

    # Configure the grid columns to be responsive
    for i in range(7):
        grid_frame.grid_columnconfigure(i, weight=1, minsize=30)
    grid_frame.grid_columnconfigure(7, weight=10, minsize=100)


    def add_new_row(customer):
        """Creates a new empty row with a checkbox and entry widgets."""
        # Calculate the next available row index
        row_index = len(entry_widgets) + 1
        
        entry_row = []
        check_var = tk.BooleanVar(value=True)
        check_button = tk.Checkbutton(grid_frame, variable=check_var)
        check_button.grid(row=row_index, column=0, sticky="nsew")
        check_vars.append(check_var)

        # Create 7 empty entry widgets for the new row
        for col_index in range(7):
            print(col_index)
            if col_index in [0, 6]:
                entry = tk.Entry(grid_frame, borderwidth=1, relief="solid")
                entry.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(entry)
            elif col_index in [2, 4]:
                entry = tk.Entry(grid_frame, borderwidth=1, relief="solid", validate='all', validatecommand=vcmd_date)
                entry.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(entry)
            elif col_index in [3, 5]:
                entry = tk.Entry(grid_frame, borderwidth=1, relief="solid", validate='all', validatecommand=vcmd_time)
                entry.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(entry)
            elif col_index == 1:
                site = tk.StringVar()
                sitedd = ttk.Combobox(grid_frame, textvariable= site)
                if customer == "Harrison St.":
                    sitedd['values'] = CUSTOMERS_SITES_NORMAL_NAMING['hst']
                elif customer == "NARENCO":
                    sitedd['values'] = CUSTOMERS_SITES_NORMAL_NAMING['nar']
                elif customer == "NCEMC":
                    sitedd['values'] = CUSTOMERS_SITES_NORMAL_NAMING['nce']
                elif customer == "Soltage":
                    sitedd['values'] = CUSTOMERS_SITES_NORMAL_NAMING['solt']
                elif customer == "Sol River":
                    sitedd['values'] = CUSTOMERS_SITES_NORMAL_NAMING['slr']
                sitedd.grid(row=row_index, column=col_index + 1, sticky="nsew")
                entry_row.append(sitedd)
        
        entry_widgets.append(entry_row)
    
    custom_data = []
    def collect_data():
        custom_data.clear()
        for entry_row, check_var in zip(entry_widgets, check_vars):
            if check_var.get():
                row_data = [entry.get() for entry in entry_row]
                custom_data.append(row_data)

    button_frame = tk.Frame(preview_window)
    button_frame.pack(fill="x", side="bottom")

    send_button = tk.Button(button_frame, text="Send Email -> Next Preview", command=lambda: [collect_data(), send_email(custom_data, preview_window, customer)], bg='lightgreen')
    send_button.pack(side="left", fill='x', expand=True)

    add_row_button = tk.Button(button_frame, text="Add Row", command=lambda: add_new_row(customer), bg='lightblue')
    add_row_button.pack(side="left", fill='x', expand=True)

    next_preview_button = tk.Button(button_frame, text="Next Preview", command=preview_window.destroy, bg='yellow')
    next_preview_button.pack(side="left", fill='x', expand=True)

    close_button = tk.Button(button_frame, text="Close Program", command=root.destroy, bg='orange')
    close_button.pack(side="left", fill='x', expand=True)


def FrameWidth(event, canvas, scrollable_frame):
    # Update the scrollregion of the canvas to match the size of the scrollable frame
    canvas.configure(scrollregion=canvas.bbox("all"))
    # Adjust the width of the scrollable frame to match the container frame's width
    canvas.itemconfig(scrollable_frame, width=event.width)

    
# Set up the tkinter window
root = tk.Tk()
root.title("NCC End of Shift Tool")
root.config(bg='gray')
vcmd_date = (root.register(legible_date_validation), '%P')
vcmd_time = (root.register(time_validation), '%P')
testingvar = tk.BooleanVar()
testingCB = tk.Checkbutton(root, text="Testing Mode", variable=testingvar, bg='gray')
testingCB.pack()

# Create a button to browse files
browse_button = tk.Button(root, text="Select Daily WO File", command=browse_files, height=3, width=40, bg='lightgreen')
browse_button.pack()
access_button = tk.Button(root, text="Site Access Reports Only", command=site_access_only, height=3, width=40, bg='yellow')
access_button.pack()
manual_button = tk.Button(root, text="Manual Reports\n(No Emaint Excel File)", command=manual_wo_reports, height=3, width=40, bg='orange')
manual_button.pack()
sSum_button = tk.Button(root, text="Shift Summary", command=shift_Summary, height=3, width=40, bg='lightblue')
sSum_button.pack()

# Run the tkinter main loop
root.mainloop()